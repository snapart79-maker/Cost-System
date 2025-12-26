"""Excel Import Service.

Excel 파일에서 데이터를 Import하는 서비스.
"""

from __future__ import annotations

import contextlib
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


@dataclass
class MaterialImportData:
    """자재 Import 데이터."""

    material_id: str
    material_name: str
    material_type: str
    unit: str
    unit_price: Decimal
    effective_date: date | None = None
    specification: str | None = None
    scrap_rate: Decimal | None = None


@dataclass
class BOMItemImportData:
    """BOM 항목 Import 데이터."""

    product_id: str
    material_id: str
    quantity: Decimal
    work_type: str
    sequence: int


@dataclass
class ProcessImportData:
    """공정 Import 데이터."""

    process_id: str
    process_name: str
    work_type: str
    labor_rate: Decimal
    machine_cost: Decimal
    efficiency: Decimal = Decimal("100.0")


@dataclass
class ImportError:
    """Import 오류 정보."""

    row: int
    column: str | None
    message: str


@dataclass
class MaterialImportResult:
    """자재 Import 결과."""

    success: bool
    total_rows: int
    imported_count: int
    error_count: int
    materials: list[MaterialImportData] = field(default_factory=list)
    errors: list[ImportError] = field(default_factory=list)


@dataclass
class BOMImportResult:
    """BOM Import 결과."""

    success: bool
    total_rows: int
    imported_count: int
    error_count: int
    bom_items: list[BOMItemImportData] = field(default_factory=list)
    errors: list[ImportError] = field(default_factory=list)


@dataclass
class ProcessImportResult:
    """공정 Import 결과."""

    success: bool
    total_rows: int
    imported_count: int
    error_count: int
    processes: list[ProcessImportData] = field(default_factory=list)
    errors: list[ImportError] = field(default_factory=list)


class ExcelImportService:
    """Excel Import 서비스."""

    def import_materials(self, file: BytesIO) -> MaterialImportResult:
        """자재 데이터 Import.

        Args:
            file: Excel 파일 바이트 스트림

        Returns:
            MaterialImportResult: Import 결과
        """
        wb = load_workbook(file, data_only=True)
        ws = wb.active

        materials: list[MaterialImportData] = []
        errors: list[ImportError] = []

        # 헤더 행 건너뛰기 (1행)
        total_rows = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 빈 행 건너뛰기
            if not row or not row[0]:
                continue

            total_rows += 1

            try:
                material = self._parse_material_row(row, row_idx)
                if material:
                    materials.append(material)
            except ValueError as e:
                errors.append(ImportError(row=row_idx, column=None, message=str(e)))

        imported_count = len(materials)
        error_count = len(errors)

        return MaterialImportResult(
            success=error_count == 0,
            total_rows=total_rows,
            imported_count=imported_count,
            error_count=error_count,
            materials=materials,
            errors=errors,
        )

    def _parse_material_row(
        self, row: tuple[Any, ...], row_idx: int
    ) -> MaterialImportData | None:
        """자재 행 파싱."""
        material_id = str(row[0]).strip() if row[0] else None
        material_name = str(row[1]).strip() if len(row) > 1 and row[1] else None
        material_type = str(row[2]).strip() if len(row) > 2 and row[2] else None
        unit = str(row[3]).strip() if len(row) > 3 and row[3] else None

        if not all([material_id, material_name, material_type, unit]):
            raise ValueError(f"Row {row_idx}: Required fields missing")

        # 단가 파싱
        try:
            unit_price = Decimal(str(row[4])) if len(row) > 4 and row[4] is not None else Decimal("0")
        except (InvalidOperation, TypeError):
            raise ValueError(f"Row {row_idx}: Invalid unit_price")

        if unit_price < 0:
            raise ValueError(f"Row {row_idx}: unit_price cannot be negative")

        # 적용일 파싱
        effective_date = None
        if len(row) > 5 and row[5]:
            if isinstance(row[5], date):
                effective_date = row[5]
            else:
                with contextlib.suppress(ValueError):
                    effective_date = date.fromisoformat(str(row[5]))

        # 규격
        specification = str(row[6]).strip() if len(row) > 6 and row[6] else None

        # SCRAP율
        scrap_rate = None
        if len(row) > 7 and row[7] is not None:
            with contextlib.suppress(InvalidOperation, TypeError):
                scrap_rate = Decimal(str(row[7]))

        return MaterialImportData(
            material_id=material_id,  # type: ignore
            material_name=material_name,  # type: ignore
            material_type=material_type,  # type: ignore
            unit=unit,  # type: ignore
            unit_price=unit_price,
            effective_date=effective_date,
            specification=specification,
            scrap_rate=scrap_rate,
        )

    def import_bom(self, file: BytesIO) -> BOMImportResult:
        """BOM 데이터 Import.

        Args:
            file: Excel 파일 바이트 스트림

        Returns:
            BOMImportResult: Import 결과
        """
        wb = load_workbook(file, data_only=True)
        ws = wb.active

        bom_items: list[BOMItemImportData] = []
        errors: list[ImportError] = []

        total_rows = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 완전히 빈 행만 건너뛰기 (일부라도 데이터가 있으면 처리)
            if not row or all(cell is None for cell in row):
                continue

            total_rows += 1

            try:
                item = self._parse_bom_row(row, row_idx)
                if item:
                    bom_items.append(item)
            except ValueError as e:
                errors.append(ImportError(row=row_idx, column=None, message=str(e)))

        imported_count = len(bom_items)
        error_count = len(errors)

        return BOMImportResult(
            success=error_count == 0,
            total_rows=total_rows,
            imported_count=imported_count,
            error_count=error_count,
            bom_items=bom_items,
            errors=errors,
        )

    def _parse_bom_row(
        self, row: tuple[Any, ...], row_idx: int
    ) -> BOMItemImportData | None:
        """BOM 행 파싱."""
        product_id = str(row[0]).strip() if row[0] else None
        material_id = str(row[1]).strip() if len(row) > 1 and row[1] else None

        if not product_id:
            raise ValueError(f"Row {row_idx}: product_id is required")
        if not material_id:
            raise ValueError(f"Row {row_idx}: material_id is required")

        # 수량 파싱
        try:
            quantity = Decimal(str(row[2])) if len(row) > 2 and row[2] is not None else Decimal("0")
        except (InvalidOperation, TypeError):
            raise ValueError(f"Row {row_idx}: Invalid quantity")

        work_type = str(row[3]).strip() if len(row) > 3 and row[3] else "IN_HOUSE"

        # 순서
        try:
            sequence = int(row[4]) if len(row) > 4 and row[4] is not None else 0
        except (ValueError, TypeError):
            sequence = 0

        return BOMItemImportData(
            product_id=product_id,
            material_id=material_id,
            quantity=quantity,
            work_type=work_type,
            sequence=sequence,
        )

    def import_processes(self, file: BytesIO) -> ProcessImportResult:
        """공정 데이터 Import.

        Args:
            file: Excel 파일 바이트 스트림

        Returns:
            ProcessImportResult: Import 결과
        """
        wb = load_workbook(file, data_only=True)
        ws = wb.active

        processes: list[ProcessImportData] = []
        errors: list[ImportError] = []

        total_rows = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue

            total_rows += 1

            try:
                process = self._parse_process_row(row, row_idx)
                if process:
                    processes.append(process)
            except ValueError as e:
                errors.append(ImportError(row=row_idx, column=None, message=str(e)))

        imported_count = len(processes)
        error_count = len(errors)

        return ProcessImportResult(
            success=error_count == 0,
            total_rows=total_rows,
            imported_count=imported_count,
            error_count=error_count,
            processes=processes,
            errors=errors,
        )

    def _parse_process_row(
        self, row: tuple[Any, ...], row_idx: int
    ) -> ProcessImportData | None:
        """공정 행 파싱."""
        process_id = str(row[0]).strip() if row[0] else None
        process_name = str(row[1]).strip() if len(row) > 1 and row[1] else None

        if not process_id or not process_name:
            raise ValueError(f"Row {row_idx}: process_id and process_name are required")

        work_type = str(row[2]).strip() if len(row) > 2 and row[2] else "IN_HOUSE"

        # 임율 파싱
        try:
            labor_rate = Decimal(str(row[3])) if len(row) > 3 and row[3] is not None else Decimal("0")
        except (InvalidOperation, TypeError):
            labor_rate = Decimal("0")

        # 기계경비 파싱
        try:
            machine_cost = Decimal(str(row[4])) if len(row) > 4 and row[4] is not None else Decimal("0")
        except (InvalidOperation, TypeError):
            machine_cost = Decimal("0")

        # 효율 파싱
        try:
            efficiency = Decimal(str(row[5])) if len(row) > 5 and row[5] is not None else Decimal("100.0")
        except (InvalidOperation, TypeError):
            efficiency = Decimal("100.0")

        return ProcessImportData(
            process_id=process_id,
            process_name=process_name,
            work_type=work_type,
            labor_rate=labor_rate,
            machine_cost=machine_cost,
            efficiency=efficiency,
        )

"""Excel Export Service.

데이터를 Excel 파일로 Export하는 서비스.
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class ExcelExportService:
    """Excel Export 서비스."""

    def __init__(self) -> None:
        """서비스 초기화."""
        # 스타일 정의
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.right_align = Alignment(horizontal="right", vertical="center")

    def export_materials(self, materials: list[dict[str, Any]]) -> BytesIO:
        """자재 목록 Export.

        Args:
            materials: 자재 데이터 목록

        Returns:
            BytesIO: Excel 파일 바이트 스트림
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Materials"

        # 헤더
        headers = [
            "material_id",
            "material_name",
            "material_type",
            "unit",
            "unit_price",
            "effective_date",
            "specification",
            "scrap_rate",
        ]
        self._write_header_row(ws, headers)

        # 데이터
        for row_idx, material in enumerate(materials, start=2):
            ws.cell(row=row_idx, column=1, value=material.get("material_id"))
            ws.cell(row=row_idx, column=2, value=material.get("material_name"))
            ws.cell(row=row_idx, column=3, value=material.get("material_type"))
            ws.cell(row=row_idx, column=4, value=material.get("unit"))
            ws.cell(
                row=row_idx,
                column=5,
                value=self._format_decimal(material.get("unit_price")),
            )
            ws.cell(
                row=row_idx,
                column=6,
                value=self._format_date(material.get("effective_date")),
            )
            ws.cell(row=row_idx, column=7, value=material.get("specification"))
            ws.cell(
                row=row_idx,
                column=8,
                value=self._format_decimal(material.get("scrap_rate")),
            )

        self._auto_column_width(ws)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def export_cost_breakdown(self, data: dict[str, Any]) -> BytesIO:
        """원가 계산서 Export.

        Args:
            data: 원가 계산 데이터

        Returns:
            BytesIO: Excel 파일 바이트 스트림
        """
        wb = Workbook()

        # 요약 시트
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._write_cost_summary(ws_summary, data)

        # 재료비 상세 시트
        if data.get("materials"):
            ws_materials = wb.create_sheet("Materials")
            self._write_materials_detail(ws_materials, data.get("materials", []))

        # 가공비 상세 시트
        if data.get("processes"):
            ws_processes = wb.create_sheet("Processes")
            self._write_processes_detail(ws_processes, data.get("processes", []))

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _write_cost_summary(self, ws: Any, data: dict[str, Any]) -> None:
        """원가 요약 작성."""
        # 제목
        ws.cell(row=1, column=1, value="원가 계산서").font = Font(bold=True, size=14)
        ws.merge_cells("A1:D1")

        # 기본 정보
        row = 3
        info_items = [
            ("완제품 ID", data.get("product_id")),
            ("완제품명", data.get("product_name")),
            ("계산일", self._format_date(data.get("calculation_date"))),
            ("버전", data.get("version")),
        ]

        for label, value in info_items:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 1

        # 원가 내역
        ws.cell(row=row, column=1, value="원가 내역").font = Font(bold=True, size=12)
        row += 1

        cost_items = [
            ("재료비", data.get("material_cost")),
            ("노무비", data.get("labor_cost")),
            ("경비", data.get("overhead_cost")),
            ("제조원가", data.get("manufacturing_cost")),
            ("재료관리비", data.get("material_management_fee")),
            ("일반관리비", data.get("general_admin_fee")),
            ("불량비", data.get("defect_cost")),
            ("이윤", data.get("profit")),
            ("구매원가", data.get("purchase_cost")),
        ]

        for label, value in cost_items:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=self._format_decimal(value))
            if label == "구매원가":
                ws.cell(row=row, column=1).font = Font(bold=True)
                ws.cell(row=row, column=2).font = Font(bold=True)
            row += 1

        self._auto_column_width(ws)

    def _write_materials_detail(self, ws: Any, materials: list[dict[str, Any]]) -> None:
        """재료비 상세 작성."""
        headers = ["material_id", "material_name", "quantity", "unit_price", "amount"]
        self._write_header_row(ws, headers)

        for row_idx, material in enumerate(materials, start=2):
            ws.cell(row=row_idx, column=1, value=material.get("material_id"))
            ws.cell(row=row_idx, column=2, value=material.get("material_name"))
            ws.cell(
                row=row_idx,
                column=3,
                value=self._format_decimal(material.get("quantity")),
            )
            ws.cell(
                row=row_idx,
                column=4,
                value=self._format_decimal(material.get("unit_price")),
            )
            ws.cell(
                row=row_idx,
                column=5,
                value=self._format_decimal(material.get("amount")),
            )

        self._auto_column_width(ws)

    def _write_processes_detail(self, ws: Any, processes: list[dict[str, Any]]) -> None:
        """가공비 상세 작성."""
        headers = ["process_id", "process_name", "labor_cost", "machine_cost"]
        self._write_header_row(ws, headers)

        for row_idx, process in enumerate(processes, start=2):
            ws.cell(row=row_idx, column=1, value=process.get("process_id"))
            ws.cell(row=row_idx, column=2, value=process.get("process_name"))
            ws.cell(
                row=row_idx,
                column=3,
                value=self._format_decimal(process.get("labor_cost")),
            )
            ws.cell(
                row=row_idx,
                column=4,
                value=self._format_decimal(process.get("machine_cost")),
            )

        self._auto_column_width(ws)

    def export_settlement(self, data: dict[str, Any]) -> BytesIO:
        """정산 내역 Export.

        Args:
            data: 정산 데이터

        Returns:
            BytesIO: Excel 파일 바이트 스트림
        """
        wb = Workbook()

        # 요약 시트
        ws_summary = wb.active
        ws_summary.title = "Settlement"
        self._write_settlement_summary(ws_summary, data)

        # 일별 상세 시트
        if data.get("daily_breakdown"):
            ws_daily = wb.create_sheet("Daily")
            self._write_daily_breakdown(ws_daily, data.get("daily_breakdown", []))

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _write_settlement_summary(self, ws: Any, data: dict[str, Any]) -> None:
        """정산 요약 작성."""
        # 제목
        ws.cell(row=1, column=1, value="정산 내역").font = Font(bold=True, size=14)
        ws.merge_cells("A1:D1")

        row = 3

        # 기본 정보
        info_items = [
            ("완제품 ID", data.get("product_id")),
            ("완제품명", data.get("product_name")),
            ("정산 기간", f"{self._format_date(data.get('period_start'))} ~ {self._format_date(data.get('period_end'))}"),
        ]

        for label, value in info_items:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 1

        # 정산 내역
        ws.cell(row=row, column=1, value="정산 내역").font = Font(bold=True, size=12)
        row += 1

        settlement_items = [
            ("변경 전 단가", data.get("before_cost")),
            ("변경 후 단가", data.get("after_cost")),
            ("단가 차이", data.get("cost_difference")),
            ("총 수량", data.get("total_quantity")),
            ("정산 금액", data.get("settlement_amount")),
        ]

        for label, value in settlement_items:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=self._format_decimal(value))
            if label == "정산 금액":
                ws.cell(row=row, column=1).font = Font(bold=True)
                ws.cell(row=row, column=2).font = Font(bold=True)
            row += 1

        self._auto_column_width(ws)

    def _write_daily_breakdown(self, ws: Any, daily_data: list[dict[str, Any]]) -> None:
        """일별 상세 작성."""
        headers = ["date", "quantity", "settlement"]
        self._write_header_row(ws, headers)

        for row_idx, day in enumerate(daily_data, start=2):
            ws.cell(row=row_idx, column=1, value=self._format_date(day.get("date")))
            ws.cell(
                row=row_idx,
                column=2,
                value=self._format_decimal(day.get("quantity")),
            )
            ws.cell(
                row=row_idx,
                column=3,
                value=self._format_decimal(day.get("settlement")),
            )

        self._auto_column_width(ws)

    def export_full_report(
        self,
        materials: list[dict[str, Any]] | None = None,
        cost_breakdown: dict[str, Any] | None = None,
        settlement: dict[str, Any] | None = None,
    ) -> BytesIO:
        """전체 보고서 Export (다중 시트).

        Args:
            materials: 자재 목록
            cost_breakdown: 원가 계산 데이터
            settlement: 정산 데이터

        Returns:
            BytesIO: Excel 파일 바이트 스트림
        """
        wb = Workbook()

        # 기본 시트 제거
        default_sheet = wb.active
        sheet_added = False

        # 자재 시트
        if materials is not None:
            ws_materials = wb.create_sheet("Materials")
            self._write_materials_sheet(ws_materials, materials)
            sheet_added = True

        # 원가 계산서 시트
        if cost_breakdown is not None:
            ws_cost = wb.create_sheet("Cost Breakdown")
            self._write_cost_summary(ws_cost, cost_breakdown)
            sheet_added = True

        # 정산 시트
        if settlement is not None:
            ws_settlement = wb.create_sheet("Settlement")
            self._write_settlement_summary(ws_settlement, settlement)
            sheet_added = True

        # 기본 시트 제거 (다른 시트가 추가된 경우)
        if sheet_added and default_sheet in wb.worksheets:
            wb.remove(default_sheet)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _write_materials_sheet(self, ws: Any, materials: list[dict[str, Any]]) -> None:
        """자재 시트 작성."""
        headers = [
            "material_id",
            "material_name",
            "material_type",
            "unit",
            "unit_price",
            "effective_date",
            "specification",
            "scrap_rate",
        ]
        self._write_header_row(ws, headers)

        for row_idx, material in enumerate(materials, start=2):
            ws.cell(row=row_idx, column=1, value=material.get("material_id"))
            ws.cell(row=row_idx, column=2, value=material.get("material_name"))
            ws.cell(row=row_idx, column=3, value=material.get("material_type"))
            ws.cell(row=row_idx, column=4, value=material.get("unit"))
            ws.cell(
                row=row_idx,
                column=5,
                value=self._format_decimal(material.get("unit_price")),
            )
            ws.cell(
                row=row_idx,
                column=6,
                value=self._format_date(material.get("effective_date")),
            )
            ws.cell(row=row_idx, column=7, value=material.get("specification"))
            ws.cell(
                row=row_idx,
                column=8,
                value=self._format_decimal(material.get("scrap_rate")),
            )

        self._auto_column_width(ws)

    def _write_header_row(self, ws: Any, headers: list[str]) -> None:
        """헤더 행 작성."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_align

    def _auto_column_width(self, ws: Any) -> None:
        """컬럼 너비 자동 조절."""
        for column_cells in ws.columns:
            max_length = 0
            column = None
            for cell in column_cells:
                try:
                    # MergedCell은 column_letter가 없으므로 건너뛰기
                    if hasattr(cell, "column_letter"):
                        if column is None:
                            column = cell.column_letter
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            if column:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

    def _format_decimal(self, value: Decimal | float | None) -> float | None:
        """Decimal 포맷팅."""
        if value is None:
            return None
        if isinstance(value, Decimal):
            return float(value)
        return value

    def _format_date(self, value: date | str | None) -> str | None:
        """날짜 포맷팅."""
        if value is None:
            return None
        if isinstance(value, date):
            return value.isoformat()
        return str(value)

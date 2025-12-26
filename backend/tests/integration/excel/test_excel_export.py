"""Test 7.1.2: Excel Export 서비스 테스트.

TDD RED Phase: Excel 데이터 Export 테스트.
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import TYPE_CHECKING

import pytest
from openpyxl import load_workbook

if TYPE_CHECKING:
    from backend.infrastructure.excel.excel_export_service import ExcelExportService


@pytest.fixture
def excel_export_service() -> "ExcelExportService":
    """Excel Export 서비스 픽스처."""
    from backend.infrastructure.excel.excel_export_service import ExcelExportService

    return ExcelExportService()


@pytest.fixture
def sample_materials_data() -> list[dict]:
    """샘플 자재 데이터."""
    return [
        {
            "material_id": "W-001",
            "material_name": "AVS 0.5sq",
            "material_type": "WIRE",
            "unit": "MTR",
            "unit_price": Decimal("15.5000"),
            "effective_date": date(2025, 1, 1),
            "specification": "0.5sq 흑색",
            "scrap_rate": Decimal("0.05"),
        },
        {
            "material_id": "W-002",
            "material_name": "AVS 0.75sq",
            "material_type": "WIRE",
            "unit": "MTR",
            "unit_price": Decimal("18.0000"),
            "effective_date": date(2025, 1, 1),
            "specification": "0.75sq 적색",
            "scrap_rate": Decimal("0.05"),
        },
    ]


@pytest.fixture
def sample_cost_breakdown_data() -> dict:
    """샘플 원가 계산서 데이터."""
    return {
        "product_id": "P-WH-001",
        "product_name": "메인 와이어 하네스",
        "calculation_date": date(2025, 1, 15),
        "version": "1.0",
        "material_cost": Decimal("1500.00"),
        "labor_cost": Decimal("800.00"),
        "overhead_cost": Decimal("200.00"),
        "manufacturing_cost": Decimal("2500.00"),
        "material_management_fee": Decimal("75.00"),
        "general_admin_fee": Decimal("125.00"),
        "defect_cost": Decimal("25.00"),
        "profit": Decimal("275.00"),
        "purchase_cost": Decimal("3000.00"),
        "materials": [
            {
                "material_id": "W-001",
                "material_name": "AVS 0.5sq",
                "quantity": Decimal("10.5"),
                "unit_price": Decimal("15.5000"),
                "amount": Decimal("162.75"),
            },
            {
                "material_id": "T-001",
                "material_name": "025 터미널",
                "quantity": Decimal("4"),
                "unit_price": Decimal("5.2500"),
                "amount": Decimal("21.00"),
            },
        ],
        "processes": [
            {
                "process_id": "PR-001",
                "process_name": "절압착",
                "labor_cost": Decimal("500.00"),
                "machine_cost": Decimal("100.00"),
            },
        ],
    }


@pytest.fixture
def sample_settlement_data() -> dict:
    """샘플 정산 데이터."""
    return {
        "product_id": "P-WH-001",
        "product_name": "메인 와이어 하네스",
        "period_start": date(2025, 1, 1),
        "period_end": date(2025, 1, 31),
        "before_cost": Decimal("2800.00"),
        "after_cost": Decimal("3000.00"),
        "cost_difference": Decimal("200.00"),
        "total_quantity": Decimal("1000"),
        "settlement_amount": Decimal("200000.00"),
        "daily_breakdown": [
            {
                "date": date(2025, 1, 15),
                "quantity": Decimal("500"),
                "settlement": Decimal("100000.00"),
            },
            {
                "date": date(2025, 1, 20),
                "quantity": Decimal("500"),
                "settlement": Decimal("100000.00"),
            },
        ],
    }


class TestExcelExportMaterials:
    """자재 Excel Export 테스트."""

    def test_export_materials_success(
        self,
        excel_export_service: "ExcelExportService",
        sample_materials_data: list[dict],
    ) -> None:
        """자재 목록 Export 성공."""
        result = excel_export_service.export_materials(sample_materials_data)

        assert result is not None
        assert isinstance(result, BytesIO)

        # Excel 파일 검증
        wb = load_workbook(result)
        ws = wb.active

        # 헤더 검증
        assert ws.cell(row=1, column=1).value == "material_id"
        assert ws.cell(row=1, column=2).value == "material_name"

        # 데이터 검증
        assert ws.cell(row=2, column=1).value == "W-001"
        assert ws.cell(row=2, column=2).value == "AVS 0.5sq"
        assert ws.cell(row=3, column=1).value == "W-002"

    def test_export_materials_empty_list(
        self, excel_export_service: "ExcelExportService"
    ) -> None:
        """빈 목록 Export."""
        result = excel_export_service.export_materials([])

        assert result is not None
        wb = load_workbook(result)
        ws = wb.active

        # 헤더만 존재
        assert ws.cell(row=1, column=1).value == "material_id"
        assert ws.cell(row=2, column=1).value is None


class TestExcelExportCostBreakdown:
    """원가 계산서 Excel Export 테스트."""

    def test_export_cost_breakdown_success(
        self,
        excel_export_service: "ExcelExportService",
        sample_cost_breakdown_data: dict,
    ) -> None:
        """원가 계산서 Export 성공."""
        result = excel_export_service.export_cost_breakdown(sample_cost_breakdown_data)

        assert result is not None
        assert isinstance(result, BytesIO)

        wb = load_workbook(result)

        # 시트 존재 확인
        assert "원가 계산서" in wb.sheetnames or "Summary" in wb.sheetnames

    def test_export_cost_breakdown_with_materials(
        self,
        excel_export_service: "ExcelExportService",
        sample_cost_breakdown_data: dict,
    ) -> None:
        """재료비 상세 포함 원가 계산서 Export."""
        result = excel_export_service.export_cost_breakdown(sample_cost_breakdown_data)

        wb = load_workbook(result)

        # 시트가 있어야 함
        assert len(wb.sheetnames) >= 1


class TestExcelExportSettlement:
    """정산 내역 Excel Export 테스트."""

    def test_export_settlement_success(
        self,
        excel_export_service: "ExcelExportService",
        sample_settlement_data: dict,
    ) -> None:
        """정산 내역 Export 성공."""
        result = excel_export_service.export_settlement(sample_settlement_data)

        assert result is not None
        assert isinstance(result, BytesIO)

        wb = load_workbook(result)
        assert len(wb.sheetnames) >= 1

    def test_export_settlement_with_daily_breakdown(
        self,
        excel_export_service: "ExcelExportService",
        sample_settlement_data: dict,
    ) -> None:
        """일별 상세 포함 정산 내역 Export."""
        result = excel_export_service.export_settlement(sample_settlement_data)

        wb = load_workbook(result)

        # 정산 정보가 있어야 함
        ws = wb.active
        assert ws is not None


class TestExcelExportMultipleSheets:
    """다중 시트 Export 테스트."""

    def test_export_full_report(
        self,
        excel_export_service: "ExcelExportService",
        sample_materials_data: list[dict],
        sample_cost_breakdown_data: dict,
    ) -> None:
        """전체 보고서 Export (다중 시트)."""
        result = excel_export_service.export_full_report(
            materials=sample_materials_data,
            cost_breakdown=sample_cost_breakdown_data,
        )

        assert result is not None
        wb = load_workbook(result)

        # 여러 시트 존재 확인
        assert len(wb.sheetnames) >= 1

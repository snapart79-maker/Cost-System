"""Test 7.1.1: Excel Import 서비스 테스트.

TDD RED Phase: Excel 데이터 Import 테스트.
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import TYPE_CHECKING

import pytest
from openpyxl import Workbook

if TYPE_CHECKING:
    from backend.infrastructure.excel.excel_import_service import ExcelImportService


@pytest.fixture
def excel_import_service() -> "ExcelImportService":
    """Excel Import 서비스 픽스처."""
    from backend.infrastructure.excel.excel_import_service import ExcelImportService

    return ExcelImportService()


@pytest.fixture
def sample_material_excel() -> BytesIO:
    """샘플 자재 Excel 파일."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Materials"

    # 헤더
    headers = [
        "material_id",
        "material_name",
        "material_type",
        "unit",
        "unit_price",
        "effective_date",
        "specification",
        "scrap_rate",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # 데이터 행
    data = [
        ["W-001", "AVS 0.5sq", "WIRE", "MTR", 15.5000, "2025-01-01", "0.5sq 흑색", 0.05],
        ["W-002", "AVS 0.75sq", "WIRE", "MTR", 18.0000, "2025-01-01", "0.75sq 적색", 0.05],
        ["T-001", "025 터미널", "TERMINAL", "EA", 5.2500, "2025-01-01", "025 암단자", 0.02],
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@pytest.fixture
def sample_bom_excel() -> BytesIO:
    """샘플 BOM Excel 파일."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"

    # 헤더
    headers = [
        "product_id",
        "material_id",
        "quantity",
        "work_type",
        "sequence",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # 데이터 행
    data = [
        ["P-WH-001", "W-001", 10.5, "IN_HOUSE", 1],
        ["P-WH-001", "W-002", 5.0, "IN_HOUSE", 2],
        ["P-WH-001", "T-001", 4, "OUTSOURCE", 3],
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@pytest.fixture
def sample_process_excel() -> BytesIO:
    """샘플 공정 Excel 파일."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Processes"

    # 헤더
    headers = [
        "process_id",
        "process_name",
        "work_type",
        "labor_rate",
        "machine_cost",
        "efficiency",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # 데이터 행
    data = [
        ["PR-001", "절압착", "IN_HOUSE", 25000.00, 5000.00, 100.0],
        ["PR-002", "테이핑", "IN_HOUSE", 20000.00, 3000.00, 95.0],
        ["PR-003", "외주 조립", "OUTSOURCE", 0, 0, 100.0],
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


class TestExcelImportMaterials:
    """자재 Excel Import 테스트."""

    def test_import_materials_success(
        self,
        excel_import_service: "ExcelImportService",
        sample_material_excel: BytesIO,
    ) -> None:
        """자재 데이터 Import 성공."""
        result = excel_import_service.import_materials(sample_material_excel)

        assert result.success is True
        assert result.total_rows == 3
        assert result.imported_count == 3
        assert result.error_count == 0
        assert len(result.materials) == 3

        # 첫 번째 자재 검증
        material = result.materials[0]
        assert material.material_id == "W-001"
        assert material.material_name == "AVS 0.5sq"
        assert material.material_type == "WIRE"
        assert material.unit == "MTR"
        assert material.unit_price == Decimal("15.5000")

    def test_import_materials_with_invalid_data(
        self, excel_import_service: "ExcelImportService"
    ) -> None:
        """잘못된 데이터가 있는 Excel Import."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Materials"

        headers = ["material_id", "material_name", "material_type", "unit", "unit_price"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # 잘못된 데이터 (unit_price가 음수)
        ws.cell(row=2, column=1, value="W-001")
        ws.cell(row=2, column=2, value="Test Wire")
        ws.cell(row=2, column=3, value="WIRE")
        ws.cell(row=2, column=4, value="MTR")
        ws.cell(row=2, column=5, value=-100)  # 음수 단가

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        result = excel_import_service.import_materials(buffer)

        assert result.error_count >= 1
        assert len(result.errors) >= 1

    def test_import_materials_empty_file(
        self, excel_import_service: "ExcelImportService"
    ) -> None:
        """빈 Excel 파일 Import."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Materials"

        # 헤더만 있는 파일
        headers = ["material_id", "material_name", "material_type", "unit", "unit_price"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        result = excel_import_service.import_materials(buffer)

        assert result.total_rows == 0
        assert result.imported_count == 0


class TestExcelImportBOM:
    """BOM Excel Import 테스트."""

    def test_import_bom_success(
        self,
        excel_import_service: "ExcelImportService",
        sample_bom_excel: BytesIO,
    ) -> None:
        """BOM 데이터 Import 성공."""
        result = excel_import_service.import_bom(sample_bom_excel)

        assert result.success is True
        assert result.total_rows == 3
        assert result.imported_count == 3

        # BOM 항목 검증
        assert len(result.bom_items) == 3
        item = result.bom_items[0]
        assert item.product_id == "P-WH-001"
        assert item.material_id == "W-001"
        assert item.quantity == Decimal("10.5")

    def test_import_bom_with_missing_product_id(
        self, excel_import_service: "ExcelImportService"
    ) -> None:
        """product_id 누락 시 오류."""
        wb = Workbook()
        ws = wb.active

        headers = ["product_id", "material_id", "quantity", "work_type", "sequence"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # product_id 누락
        ws.cell(row=2, column=1, value="")  # 빈 값
        ws.cell(row=2, column=2, value="W-001")
        ws.cell(row=2, column=3, value=10)
        ws.cell(row=2, column=4, value="IN_HOUSE")
        ws.cell(row=2, column=5, value=1)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        result = excel_import_service.import_bom(buffer)

        assert result.error_count >= 1


class TestExcelImportProcesses:
    """공정 Excel Import 테스트."""

    def test_import_processes_success(
        self,
        excel_import_service: "ExcelImportService",
        sample_process_excel: BytesIO,
    ) -> None:
        """공정 데이터 Import 성공."""
        result = excel_import_service.import_processes(sample_process_excel)

        assert result.success is True
        assert result.total_rows == 3
        assert result.imported_count == 3

        # 공정 검증
        process = result.processes[0]
        assert process.process_id == "PR-001"
        assert process.process_name == "절압착"
        assert process.labor_rate == Decimal("25000.00")
